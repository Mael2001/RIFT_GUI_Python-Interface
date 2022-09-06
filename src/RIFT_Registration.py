#Base Libraries
from cgitb import text
from cmath import nan
from logging import PlaceHolder
from wsgiref import validate
import pandas as pd
import os
from openpyxl import *
from datetime import datetime
import time
import webbrowser
import threading

#Graphic Libraries
import tkinter.messagebox
import tkinter
import customtkinter
import easygui

#Image Libraries
from PIL import Image, ImageTk

#SIZE
COMPLETE_WIDTH = 700
COMPLETE_HEIGHT = 800

#Working Directories
CURRENT_DIRECTORY = os.path.dirname(os.path.realpath(__file__))
FILE_NAME=".\dump\\RIFT.xlsx"
FILE_PATH = os.path.join(CURRENT_DIRECTORY,FILE_NAME)

class Registration(customtkinter.CTkToplevel):

    APP_WIDTH = COMPLETE_WIDTH
    APP_HEIGHT = COMPLETE_HEIGHT
    current_team=[]
    stop = False
    fish_types = {
        "Billfish":[
            "Blue Marlin",
            "White Marlin",
            "Sailfish & Spearfish"
        ],
        "Rodeo":[
            "Wahoo",
            "Dolphin",
            "Tuna",
            "Kingfish"
        ],
        "Junior":[
            "Barracuda",
            "Mackerel",
            "King Fish",
            "Wahoo",
            "Dolphin MahiMahi",
            "Tuna"
        ],
        "Kids":[
            "Barracuda",
            "Mackerel",
            "King Fish",
            "Wahoo",
            "Dolphin MahiMahi",
            "Tuna"
        ],
        "Women":[
            "Barracuda",
            "Mackerel",
            "King Fish",
            "Wahoo",
            "Dolphin MahiMahi",
            "Tuna"
        ],
        "None":[
            "None"
        ]
    }
    fish_prices = {
            "Blue Marlin":500,
            "White Marlin":300,
            "Sailfish & Spearfish":200,

            "Wahoo":0,
            "Dolphin":0,
            "Tuna":0,
            "Kingfish":0,

            "Barracuda":0,
            "Mackerel":0,
            "King Fish":0,
            "Wahoo":0,
            "MahiMahi":0,
            "Dolphin":0,
            "Tuna":0
            }

    def __init__(self,parent):
        super().__init__(parent)
        WIDTH = self.winfo_screenwidth()
        HEIGHT = self.winfo_screenwidth()

        app_center_coordinate_x = (WIDTH/2) - (Registration.APP_WIDTH )
        app_center_coordinate_y = (HEIGHT/2) - (Registration.APP_HEIGHT)

        self.title("Plataforma RIFT")
        self.geometry(f"{Registration.APP_WIDTH}x{Registration.APP_HEIGHT}+{int(app_center_coordinate_x)}+{int(app_center_coordinate_y)}")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)  # call .on_closing() when app gets closed
        self.configure(fg_color=("#189FE7"))
        image = Image.open("./images/BckImage.jpeg").resize((self.APP_WIDTH, self.APP_HEIGHT))
        self.bg_image = ImageTk.PhotoImage(image)

        self.image_label = tkinter.Label(master=self, image=self.bg_image)
        self.image_label.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

        #GUI TITLE
        self.homeLabel = customtkinter.CTkLabel(master=self,
                                                    corner_radius=7,
                                                    height=40,
                                                    fg_color=("white", "grey38"),  # <- custom tuple-color
                                                    text_font=("Arial",14),
                                                    text="Hookup Registration")
        self.homeLabel.grid(row=0,column=0,columnspan=2,padx=15,pady=10)

        #BASE FRAME
        self.frame_home = customtkinter.CTkFrame(master=self,
                                                    width= Registration.APP_WIDTH,
                                                    height= Registration.APP_HEIGHT,
                                                    fg_color="black",
                                                    corner_radius=10)
        self.frame_home.grid(row=1,column=0,padx=15, pady=20,sticky="nsew")


        #BOAT NAME
        self.boat_name = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Enter Boat ID: ")
        self.boat_name.grid(row=0,column=0,padx=10,pady=20)

        #BOAT INPUT
        self.boat_input= customtkinter.CTkEntry(master=self.frame_home,
                                            width=Registration.APP_WIDTH/2+50,
                                            placeholder_text="Boat ID")
        self.boat_input.grid(row=0,column=1, columnspan=2, pady=20, padx=20, sticky="we")

        #SEARCH BUTTON
        self.search_button= customtkinter.CTkButton(master=self.frame_home,
                                            width=Registration.APP_WIDTH/2-50,
                                            text="Search",
                                            command=self.search)
        self.search_button.grid(row=1,column=0,columnspan=3,pady=20, padx=20, sticky="we")

        #SEARCH FRAME
        self.search_frame = customtkinter.CTkFrame(master=self.frame_home,
                                                    width= Registration.APP_WIDTH,
                                                    height= Registration.APP_HEIGHT,
                                                    fg_color="white",
                                                    corner_radius=10)
        self.search_frame.grid(row=2,column=0,columnspan=4,padx=15, pady=20,sticky="nsew")

        #BOAT NAME
        self.search_boat_name = customtkinter.CTkEntry(master=self.search_frame,
                                                    corner_radius=7,
                                                    width= Registration.APP_WIDTH/2-50,
                                                    placeholder_text="Boat Name",
                                                    text="Boat")
        self.search_boat_name.grid(row=0,column=0,columnspan=2,padx=10,pady=20)
        self.search_boat_name.configure(state=tkinter.DISABLED)

        #CAPTAIN NAME
        self.search_captain_name = customtkinter.CTkEntry(master=self.search_frame,
                                                    corner_radius=7,
                                                    width= Registration.APP_WIDTH/2-50,
                                                    placeholder_text="Captain Name",
                                                    text="Captain")
        self.search_captain_name.grid(row=0,column=3,columnspan=2,padx=10,pady=20)
        self.search_captain_name.configure(state=tkinter.DISABLED)

        #CATEGORY SELECT LABEL
        self.category_combobox_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Select Category: ")
        self.category_combobox_label.grid(row=3,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #CATEGORY SELECT
        self.category_combobox = customtkinter.CTkOptionMenu(master=self.frame_home,
                                                    values=["Billfish", "Rodeo", "Junior","Kids", "Women", "None"])
        self.category_combobox.grid(row=3,column=1,columnspan=2,padx=15, pady=20,sticky="nsew")
        self.category_combobox.set("None")
        self.category_combobox.configure(state=tkinter.DISABLED)


        #FISH SELECT LABEL
        self.fish_combobox_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Select Fish: ")
        self.fish_combobox_label.grid(row=4,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #FISH SELECT
        self.fish_combobox = customtkinter.CTkOptionMenu(master=self.frame_home,
                                                    values=["None"])
        self.fish_combobox.grid(row=4,column=1,columnspan=2,padx=15, pady=20,sticky="nsew")
        self.fish_combobox.set("None")
        self.fish_combobox.configure(state=tkinter.DISABLED)
        self.category_combobox.configure(command=self.update_fish_list)

        #HOOKUP TIME LABEL
        self.hookup_time_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Hookup Time: ")
        self.hookup_time_label.grid(row=5,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #HOOKUP TIME
        self.hookup_time = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text=datetime.now().strftime("%H:%M:%S"),
                                                    width= Registration.APP_WIDTH/2-50)
        self.hookup_time.grid(row=5,column=1,columnspan=1,padx=15, pady=20,sticky="nsew")

        #REGISTER HOOKUP
        self.hookup_button = customtkinter.CTkButton(master=self.frame_home,
                                            width=Registration.APP_WIDTH/2-50,
                                            text="Register Hookup",
                                            command=self.register_hookup)
        self.hookup_button.grid(row=6,column=0,columnspan=3,padx=15, pady=20,sticky="nsew")
        self.hookup_button.configure(state=tkinter.DISABLED)

        #RELEASE TIME LABEL
        self.release_time_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Release Time: ")
        self.release_time_label.grid(row=7,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #RELEASE TIME
        self.release_time = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text=datetime.now().strftime("%H:%M:%S"),
                                                    width= Registration.APP_WIDTH/2-50)
        self.release_time.grid(row=7,column=1,columnspan=1,padx=15, pady=20,sticky="nsew")

        #CLEAN RELEASE CHECK
        self.clean_release=customtkinter.CTkCheckBox(master=self.frame_home,
                                                    text="Clean Release")
        self.clean_release.grid(row=7,column=2,columnspan=1,padx=15, pady=20,sticky="nsew")
        self.clean_release.configure(state=tkinter.DISABLED)

        #DISCARD BUTTON
        self.discard_button = customtkinter.CTkButton(master=self.frame_home,
                                            text="Discard",
                                            command=self.discard)
        self.discard_button.grid(row=8,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #CHECK XLSX BUTTON
        self.discard_button = customtkinter.CTkButton(master=self.frame_home,
                                            text="View Worksheet",
                                            command=self.view)
        self.discard_button.grid(row=8,column=1,columnspan=1,padx=15, pady=20,sticky="nsew")

        #REGISTER RELEASE
        self.release_button = customtkinter.CTkButton(master=self.frame_home,
                                            text="Register Release",
                                            command=self.register_release)
        self.release_button.grid(row=8,column=2,columnspan=1,padx=15, pady=20,sticky="nsew")
        self.release_button.configure(state=tkinter.DISABLED)

        clock_thread = threading.Thread(target=self.updateClock)
        clock_thread.start()

    #Update Available Fishes
    def update_fish_list(self,value):
        self.fish_combobox.configure(values=self.fish_types[self.category_combobox.get()])

    #Register hookup
    def register_hookup(self):
        if(self.validate_inputs()):
            sheet_name=self.category_combobox.get()
            wb = load_workbook(filename=FILE_NAME)
            ws = wb[sheet_name]
            if(ws.max_row < 2):
                headers = [
                    "HOOKUP_ID",
                    "ID",
                    "BOAT_NAME",
                    "CAPTAIN_NAME",
                    "HOOKUP_TIME",
                    "FISH",
                    "RELEASE_TIME",
                    "CLEAN_RELEASE",
                    "POINTS"
                    ]
                self.write_to_sheet(FILE_NAME,sheet_name,headers)
            data = [
                ws.max_row,
                self.current_team['ID'],
                self.current_team['BOAT_NAME'],
                self.current_team['CAPTAIN_NAME'],
                datetime.now().strftime("%H:%M:%S"),
                self.fish_combobox.get(),
                "None",
                "No",
                self.fish_prices[self.fish_combobox.get()]
            ]
            print(f"Wrote to {FILE_NAME}, sheet {sheet_name}")
            self.write_to_sheet(FILE_NAME,sheet_name,data)
            self.trigger_message(f"Hookup Registered at {FILE_NAME}, sheet {sheet_name}","Hookup Registration")
            self.fish_prices[self.fish_combobox.get()] = 0
        else:
            self.trigger_error("Please Fix Errors in the form","Unable to Continue")
            return

    #Validation for all inputs
    def validate_inputs(self):
        if(self.category_combobox.get() == "None" or self.current_team[f"{self.category_combobox.get().upper()}_PRESENT"] == "Not Participating"):
            self.trigger_error("Team is not participating in this Category", "Input error")
            return False
        if(self.fish_combobox.get() == "None"):
            self.trigger_error("Invalid Fish Selected", "Input error")
            return False
        if(self.fish_prices[self.fish_combobox.get()] == 0):
            self.fish_prices[self.fish_combobox.get()] = int(easygui.enterbox("What's the fish Weight"))
            if(self.fish_prices[self.fish_combobox.get()] < 20):
                self.trigger_error("Fish must weigh atleast 20 pounds", "Input error")
                self.fish_prices[self.fish_combobox.get()] = 0
                return False
        return True
    #Validation for Release
    def validate_release(self):
        if(isinstance(self.current_team, pd.Series) == False):
            self.trigger_error("No Team Searched", "Missing Action")
            return False
        if(self.category_combobox.get() == "None" or self.current_team[f"{self.category_combobox.get().upper()}_PRESENT"] == "Not Participating"):
            self.trigger_error("No Category Selected", "Missing Input")
            return False
        if(self.boat_input.get() == ""):
            self.trigger_error("No Boat Id Given", "Missing Input")
            return False
        return True
    #Register release
    def register_release(self):
        if(self.validate_release()):
            sheetName=self.category_combobox.get()
            hookup_result = self.search_hookup(sheetName)
            self.rewrite_row(FILE_NAME,sheetName,hookup_result)
            print(f"Wrote to {FILE_NAME}, sheet {sheetName}")
            self.trigger_message(f"Release Registered at {FILE_NAME}, sheet {sheetName}","Hookup Registration")
        else:
            self.trigger_error("Please Fix Errors in the form","Unable to Continue")
            return

    #Search Hookup Information
    def search_hookup(self,sheetName):
        excel_data = pd.read_excel(FILE_NAME,sheet_name=sheetName)
        boat_id = self.boat_input.get()
        hookup = excel_data.loc[excel_data['ID']==int(boat_id)]
        if(len(hookup)>1):
            return hookup.iloc[-1]
        return hookup

    #Write Release Information
    def rewrite_row(self,path,sheetName,hookup_result):
        if(hookup_result.empty):
            self.trigger_error("Participant is not registered for this category, please ensure correct category was selected", "Participant not found")
        wb = load_workbook(path)
        ws = wb[sheetName]
        row = (hookup_result['HOOKUP_ID'])+1
        row = row.values[0]
        print(hookup_result)
        RELEASE_TIME = ws.cell(row=row,column=7)
        RELEASE_TIME.value = datetime.now().strftime("%H:%M:%S")
        CLEAN_RELEASE = ws.cell(row=row,column=8)
        CLEAN_RELEASE.value = "Yes" if self.clean_release.get() else "No"
        POINTS = ws.cell(row=row,column=9)
        POINTS.value = hookup_result['POINTS'].values[0] + 50 if self.clean_release.get() else hookup_result['POINTS'].values[0]
        wb.save(path)

    #View XLSX file
    def view(self):
        webbrowser.open(FILE_NAME)
    #Discard all changes
    def discard(self):
        #Resetting Search Results
        self.search_boat_name.configure(state=tkinter.NORMAL)
        self.search_captain_name.configure(state=tkinter.NORMAL)
        self.search_boat_name.configure(placeholder_text="Boat Name")
        self.search_captain_name.configure(placeholder_text="Captain Name")
        self.search_boat_name.configure(state=tkinter.DISABLED)
        self.search_captain_name.configure(state=tkinter.DISABLED)
        #Refreshing Clocks
        self.hookup_time.configure(text=datetime.now().strftime("%H:%M:%S"))
        self.release_time.configure(text=datetime.now().strftime("%H:%M:%S"))
        #Resetting default values
        self.boat_input.configure(text="")
        self.category_combobox.set("None")
        self.fish_combobox.set("None")
        self.clean_release.deselect()
        #Disabling all controls when refreshing form
        self.category_combobox.configure(state=tkinter.DISABLED)
        self.fish_combobox.configure(state=tkinter.DISABLED)
        self.hookup_button.configure(state=tkinter.DISABLED)
        self.clean_release.configure(state=tkinter.DISABLED)
        self.release_button.configure(state=tkinter.DISABLED)
        return True
    #Search ID
    def search(self):
        excel_data = pd.read_excel(FILE_NAME,sheet_name="Inscription")
        boat_id = self.boat_input.get()
        for result in (excel_data.isin([int(boat_id)]).any()):
            if result:
                #Getting team information
                self.current_team = excel_data.iloc[int(boat_id)-1]
                #Writing Search Results
                self.search_boat_name.configure(state=tkinter.NORMAL)
                self.search_captain_name.configure(state=tkinter.NORMAL)
                self.search_boat_name.configure(placeholder_text=self.current_team['BOAT_NAME'])
                self.search_captain_name.configure(placeholder_text=self.current_team['CAPTAIN_NAME'])
                self.search_boat_name.configure(state=tkinter.DISABLED)
                self.search_captain_name.configure(state=tkinter.DISABLED)
                #Enabling all controls when Successfully found ID
                self.category_combobox.configure(state=tkinter.NORMAL)
                self.fish_combobox.configure(state=tkinter.NORMAL)
                self.hookup_button.configure(state=tkinter.NORMAL)
                self.clean_release.configure(state=tkinter.NORMAL)
                self.release_button.configure(state=tkinter.NORMAL)
                return
        self.trigger_error(f"No Inscription was found matching {boat_id}","No Register Found")
    #Writing to File
    def write_to_sheet(self,path,sheetName,data):
        wb = load_workbook(path)
        ws = wb[sheetName]
        ws.append(data)
        wb.save(path)
    #Update HookupTime
    def updateClock(self):
        while self.stop == False:
            self.hookup_time.configure(text=datetime.now().strftime("%H:%M:%S"))
            self.release_time.configure(text=datetime.now().strftime("%H:%M:%S"))
            time.sleep(1)
    #Information Box Trigger
    def trigger_message(self,Msg,Title):
        tkinter.messagebox.showinfo(title=Title, message=Msg)
    #Error Box Trigger
    def trigger_error(self,ErrorMsg,Error):
        tkinter.messagebox.showerror(title=Error, message=ErrorMsg)
    #Close Program
    def on_closing(self, event=0):
        self.stop = True
        time.sleep(0.1)
        self.destroy()

if __name__ == "__main__":
    app = Registration()
    app.resizable(False,False)
    app.set_BckImage()
    app.mainloop()