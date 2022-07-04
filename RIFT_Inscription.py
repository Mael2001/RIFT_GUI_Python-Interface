#Base Libraries
import pandas as pd
import os
from openpyxl import *
import phonenumbers
import re

#Graphic Libraries
import tkinter.messagebox
import tkinter
import customtkinter

#Image Libraries

#SIZE
COMPLETE_WIDTH = 600
COMPLETE_HEIGHT = 600

#Working Directories
CURRENT_DIRECTORY = os.path.dirname(os.path.realpath(__file__))
FILE_NAME=".\dump\\test.xlsx"
FILE_PATH = os.path.join(CURRENT_DIRECTORY,FILE_NAME)

#Base Settings
customtkinter.set_appearance_mode("system")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green



class Inscription(customtkinter.CTk):

    APP_WIDTH = COMPLETE_WIDTH
    APP_HEIGHT = COMPLETE_HEIGHT

    def __init__(self):
        super().__init__()
        WIDTH = self.winfo_screenwidth()
        HEIGHT = self.winfo_screenwidth()

        app_center_coordinate_x = (WIDTH/2) - (Inscription.APP_WIDTH )
        app_center_coordinate_y = (HEIGHT/2) - (Inscription.APP_HEIGHT)

        self.title("Plataforma RIFT")
        self.geometry(f"{Inscription.APP_WIDTH}x{Inscription.APP_HEIGHT}+{int(app_center_coordinate_x)}+{int(app_center_coordinate_y)}")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)  # call .on_closing() when app gets closed
        self.configure(fg_color=("#189FE7"))
        #Titulo
        #bckImage = Image.open(os.path.join(CURRENT_DIRECTORY,"/images/bg.png"))
        #bckImage = Image.open("./images/bg.png")
        #bckImage_resized = bckImage.resize((WIDTH,HEIGHT))
        #bckImage = ImageTk.PhotoImage(bckImage_resized)


        #GUI BACKGROUND
        #homeBckImage = customtkinter.CTkLabel(master=self,
        #                                            corner_radius=7,
        #                                            image=bckImage)
        #homeBckImage.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)
        #GUI TITLE
        self.homeLabel = customtkinter.CTkLabel(master=self,
                                                    corner_radius=7,
                                                    height=40,
                                                    fg_color=("white", "grey38"),  # <- custom tuple-color
                                                    text_font=("Arial",14),
                                                    text="Inscription")
        self.homeLabel.grid(row=0,column=0,columnspan=2,padx=15,pady=10)

        #BASE FRAME
        self.frame_home = customtkinter.CTkFrame(master=self,
                                                    width= Inscription.APP_WIDTH,
                                                    height= Inscription.APP_HEIGHT,
                                                    fg_color="black",
                                                    corner_radius=10)
        self.frame_home.grid(row=1,column=0,padx=15, pady=20,sticky="nsew")

        #BOAT NAME
        self.boat_name = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Enter Boat Name: ")
        self.boat_name.grid(row=0,column=0,padx=10,pady=20)

        #BOAT INPUT
        self.boat_input= customtkinter.CTkEntry(master=self.frame_home,
                                            width=Inscription.APP_WIDTH/2+50,
                                            placeholder_text="Boat Name")
        self.boat_input.grid(row=0,column=1, columnspan=2, pady=20, padx=20, sticky="we")

        #CAPTAIN NAME
        self.captain_name = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Enter Captain Name: ")
        self.captain_name.grid(row=1,column=0,padx=10,pady=20)

        #CAPTAIN INPUT
        self.captain_input= customtkinter.CTkEntry(master=self.frame_home,
                                            width=Inscription.APP_WIDTH/2+50,
                                            placeholder_text="Captain Name")
        self.captain_input.grid(row=1,column=1, columnspan=2, pady=20, padx=20, sticky="we")


        #PHONE NUMBER LABEL
        self.phone_number_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Enter Phone Number: ")
        self.phone_number_label.grid(row=2,column=0,padx=10,pady=20)

        #PHONE NUMBER INPUT
        self.phone_number_input= customtkinter.CTkEntry(master=self.frame_home,
                                            width=Inscription.APP_WIDTH/2+50,
                                            placeholder_text="000-000-000")
        self.phone_number_input.grid(row=2,column=1, columnspan=2, pady=20, padx=20, sticky="we")

        #EMAIL NUMBER LABEL
        self.email_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Enter Email: ")
        self.email_label.grid(row=3,column=0,padx=10,pady=20)

        #EMAIL NUMBER INPUT
        self.email_input= customtkinter.CTkEntry(master=self.frame_home,
                                            width=Inscription.APP_WIDTH/2+50,
                                            placeholder_text="email@test.com")
        self.email_input.grid(row=3,column=1, columnspan=2, pady=20, padx=20, sticky="we")

        #SECTION LABEL
        categories_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    width=Inscription.APP_WIDTH-50,
                                                    fg_color=("white", "grey38"),  # <- custom tuple-color
                                                    text_font=("Arial",14),
                                                    text="Categories")
        categories_label.grid(row=4,column=0,columnspan=3,padx=15,pady=10)

        #BILLFISH CATEGORY
        self.billfish_category = customtkinter.CTkCheckBox(master=self.frame_home,
                                                    text="Billfish")
        self.billfish_category.grid(row=5,column=0,columnspan=1,padx=15,pady=10)

        #RODEO CATEGORY
        self.rodeo_category = customtkinter.CTkCheckBox(master=self.frame_home,
                                                    text="Rodeo")
        self.rodeo_category.grid(row=5,column=1,columnspan=1,padx=15,pady=10)

        #JUNIOR CATEGORY
        self.junior_category = customtkinter.CTkCheckBox(master=self.frame_home,
                                                    text="Junior")
        self.junior_category.grid(row=5,column=2,columnspan=1,padx=15,pady=10)

        #KIDS CATEGORY
        self.kids_category = customtkinter.CTkCheckBox(master=self.frame_home,
                                                    text="Kids")
        self.kids_category.grid(row=6,column=0,columnspan=2,padx=15,pady=10)

        #WOMEN CATEGORY
        self.women_category = customtkinter.CTkCheckBox(master=self.frame_home,
                                                    text="Women")
        self.women_category.grid(row=6,column=1,columnspan=2,padx=15,pady=10)


        #REGISTER BUTTON
        self.register_button = customtkinter.CTkButton(master=self.frame_home,
                                                    text="Register",
                                                    width=Inscription.APP_WIDTH-50,
                                                    text_color="white",
                                                    command=self.register)
        self.register_button.grid(row=7,column=0,columnspan=4,padx=15,pady=10)

    #Dump Values to XLSX file
    def register(self):
        #Validating File Existence and creating it
        if (os.path.exists(FILE_NAME) == False):
            print(f"Creating file at {FILE_NAME}")
            self.create_workbook(FILE_NAME)
            headers = [
                "ID",
                "BOAT_NAME",
                "CAPTAIN_NAME",
                "PHONE_NUMBER",
                "EMAIL",
                "BILLFISH_PRESENT",
                "RODEO_PRESENT",
                "JUNIOR_PRESENT",
                "KIDS_PRESENT",
                "WOMEN_PRESENT"
                ]
            self.write_to_sheet(FILE_NAME,"Inscription",headers)
        #Writing data to File
        wb = load_workbook(filename=FILE_NAME)
        ws = wb["Inscription"]
        data = [
            ws.max_row,
            self.boat_input.get(),
            self.captain_input.get(),
            self.phone_number_input.get(),
            "Participating" if self.billfish_category.get() else "Not Participating",
            "Participating" if self.rodeo_category.get() else "Not Participating",
            "Participating" if self.junior_category.get() else "Not Participating",
            "Participating" if self.kids_category.get() else "Not Participating",
            "Participating" if self.women_category.get() else "Not Participating"
        ]
        if(self.validate_fields()):
            self.write_to_sheet(FILE_NAME,"Inscription",data)
            print(f"Wrote to {FILE_NAME}")
        else:
            print("Invalid inputs detected")

    #Creating XLSX File
    def create_workbook(self,path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inscription"
        workbook.save(path)
    #Creating New Sheet
    def create_sheets(self,path):
        workbook = Workbook()
        workbook.create_sheet("Billfish")
        workbook.create_sheet("Rodeo")
        workbook.create_sheet("Junior")
        workbook.create_sheet("Kids")
        workbook.create_sheet("Women")
        workbook.save(path)
    #Writing to File
    def write_to_sheet(self,path,sheetName,data):
        wb = load_workbook(path)
        ws = wb[sheetName]
        ws.append(data)
        wb.save(path)
    #Validate input fields
    def validate_fields(self):
        #Validating Boat Input
        if(self.validate_field_excel(FILE_NAME,"Inscription",self.boat_input.get())== False or not self.boat_input.get()):
            self.trigger_error("Boat name Already Exist","Invalid Boat")
            return False
        #Validating Captain Input
        if(self.validate_field_excel(FILE_NAME,"Inscription",self.captain_input.get())== False or not self.captain_input.get()):
            self.trigger_error("Captain name Already Exist","Invalid Captain")
            return False
        #Validating Phone Number Input
        if(self.validate_phone_number(self.phone_number_input.get())== False or not self.phone_number_input.get()):
            self.trigger_error("Phone Number is not valid Input","Invalid Phone")
            return False
        if(self.validate_email(self.email_input.get())== False or not self.email_input.get()):
            self.trigger_error("Email is not valid Input","Invalid Email")
            return False
        if(self.validate_categories()==False):
            self.trigger_error("No Categories have been selected","Error")
            return False
        return True
    #Validate repeated values in file
    def validate_field_excel(self,path,sheet,value):
        excel_data = pd.read_excel(path,sheet_name=sheet)
        for result in (excel_data.isin([value]).any()):
            if result:
                return False
        return True
    #Validate Phone Number
    def validate_phone_number(self,value):
        regex = "^(\+\d{1,2}\s)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}$"
        if re.match(regex,value):
            return True
        return False
    #Validate Email
    def validate_email(self,value):
        regex = "^[a-zA-Z0-9-_]+@[a-zA-Z0-9]+\.[a-z]{1,3}$"
        if re.match(regex,value):
            return True
        return False
    #Validate Categories
    def validate_categories(self):
        if(self.billfish_category.get()==0
        and self.rodeo_category.get()==0
        and self.junior_category.get()==0
        and self.kids_category.get()==0
        and self.women_category.get()==0):
            return False
        return True
    #Error Box Trigger
    def trigger_error(self,ErrorMsg,Error):
        tkinter.messagebox.showerror(title=Error, message=ErrorMsg)
    #Close Program
    def on_closing(self, event=0):
        self.destroy()
