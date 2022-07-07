import time
import tkinter as tk
from tkinter import Frame, Image, Label, Radiobutton, Text, filedialog
from tkinter.constants import CENTER, FALSE, LEFT, NONE, RIGHT, TRUE
from tkinter.font import Font
import cv2
from tkinter import messagebox
from tkinter import font
import os
from stat import S_IREAD, S_IWRITE
import datetime as dt
from tkinter import ttk
import numpy as np
from openpyxl.descriptors.base import Length
from openpyxl.workbook import workbook
import pandas as pd
import matplotlib as plt
import openpyxl
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox

root = tk.Tk()
root.title('XXII Roatan International Fishing Tournament')

HEIGHT = 700
WIDTH = 1600

canvas = tk.Canvas(root, height=HEIGHT, width=WIDTH)
canvas.pack()

def clock():
    t=time.strftime('%I:%M:%S',time.localtime())
    if t!='':
        dt_label.config(text=t,font='times 12')

    root.after(100,clock)

    return(t)

def isChecked_C1():

  if CheckVar1.get() == 1:
        estado = TRUE
  else:
         estado = FALSE

  return(estado)

def isChecked_C2():

   if CheckVar2.get() == 1:
        estado = TRUE
   else:
         estado = FALSE

   return(estado)

def isChecked_C3():

  if CheckVar3.get() == 1:
        estado = TRUE
  else:
         estado = FALSE

  return(estado)

def isChecked_C4():

   if CheckVar4.get() == 1:
        estado = TRUE
   else:
         estado = FALSE

   return(estado)

def isChecked_C5():

   if CheckVar5.get() == 1:
        estado = TRUE
   else:
         estado = FALSE

   return(estado)

#Set Background
bg_image = tk.PhotoImage(file='./images/bg2.png')
bg_label = tk.Label(root, image=bg_image)
bg_label.place(relwidth=1, relheight=1)

#Design config for inscription
frame = tk.Frame(root, bg='white')
frame.place(relx=0.03, rely=0.05, relwidth=0.45, relheight = 0.4)

i_image = tk.PhotoImage(file='./images/inscrip.png')
i_label = tk.Label(frame, image=i_image, bg='white')
i_label.place(relx=0.85, rely=0.2)

bg_label = tk.Label(frame, text='Inscription', font="Arialnova 25")
bg_label.place(relwidth=1, relheight=0.15)

embarcation = tk.Entry(frame, cursor="arrow", font="times 14")
embarcation.place(relx=0.55, rely=0.22, relheight=0.095, relwidth=0.2)
embarc_label = tk.Label(frame, text='Enter Boat Name: ', font="times 14")
embarc_label.place(relx=0.25, rely=0.22)

captain = tk.Entry(frame, cursor="arrow", font="times 14")
captain .place(relx=0.55, rely=0.32, relheight=0.095, relwidth=0.2)
captain_label = tk.Label(frame, text='Enter Captain Name: ', font="times 14")
captain_label.place(relx=0.25, rely=0.32)

cat_label = tk.Label(frame, text='Categories', font="times 16", anchor='n')
cat_label.place(relx=0.43, rely=0.54)

number = tk.Entry(frame, cursor="arrow", font="times 14")
number.place(relx=0.55, rely=0.42, relheight=0.095, relwidth=0.2)
num_label = tk.Label(frame, text='Phone Number: ', font="times 14")
num_label.place(relx=0.25, rely=0.42)

font_text_ins = "times 12"

CheckVar1 = tk.IntVar()
CheckVar2 = tk.IntVar()
C1 = tk.Checkbutton(frame, text = "Billfish", command=isChecked_C1, bg = "white", font=font_text_ins, variable = CheckVar1, onvalue = 1, offvalue = 0, height=2, width = 20, anchor='w')
C2 = tk.Checkbutton(frame, text = "Rodeo",command=isChecked_C2, bg = "white", font=font_text_ins, variable = CheckVar2, onvalue = 1, offvalue = 0, height=3, width = 20, anchor='w')
C1.place(relx=0.15, rely=0.64)
C2.place(relx=0.3, rely=0.64)
CheckVar3 = tk.IntVar()
CheckVar4 = tk.IntVar()
C3 = tk.Checkbutton(frame, text = "Junior", command=isChecked_C3, bg = "white",font=font_text_ins, variable = CheckVar3, onvalue = 1, offvalue = 0, height=2, width = 20, anchor='w')
C4 = tk.Checkbutton(frame, text = "Kids",command=isChecked_C4, bg = "white",font=font_text_ins, variable = CheckVar4, onvalue = 1, offvalue = 0, height=3, width = 20, anchor='w')
C3.place(relx=0.45, rely=0.64)
C4.place(relx=0.60, rely=0.64)
CheckVar5 = tk.IntVar()
C5 = tk.Checkbutton(frame, text = "Women", command=isChecked_C5, bg = "white",font=font_text_ins, variable = CheckVar5, onvalue = 1, offvalue = 0, height=2, width = 20, anchor='w')
C5.place(relx=0.75, rely=0.64)

def focusNext(widget):
 widget.tk_focusNext().focus_set()
 return 'break'

def focusPrev(widget):
 widget.tk_focusPrev().focus_set()
 return 'break'

def Iterador_billfish():

 os.chmod("XXII_RIFT.xlsx", S_IWRITE)

 wb = load_workbook('XXII_RIFT.xlsx')
 ws = wb['Billfish']

 lenght = len(ws['A'])
 print(lenght)

 return(lenght)

def get_IDnum():

 os.chmod("XXII_RIFT.xlsx", S_IWRITE)

 wb = load_workbook('XXII_RIFT.xlsx')
 ws = wb['Inscription']

 lenght = len(ws['A'])
 print(lenght)

 return(lenght)

def success_inscrip():

      messagebox.showinfo("Vessels", "Vessel Registration Succesful")

def failed_inscrip():

      messagebox.showwarning("Vessels", "Vessel Registration Unsuccesful, all fields must be populated")

def wrong_num_length():

      messagebox.showwarning("Vessels", "Phone number must be 8 digits")

def register_inscrip():

  os.chmod("XXII_RIFT.xlsx", S_IWRITE)

  ID = get_IDnum()

  wb = load_workbook('XXII_RIFT.xlsx')
  ws = wb['Inscription']

  if isChecked_C1() == True:

        bf = "Billfish"

  else:

        bf = ''

  if isChecked_C2() == True:

        rd = "Rodeo"


  else:

        rd = ''

  if isChecked_C3() == True:

        jr = "Junior"


  else:
        jr = ''

  if isChecked_C4() == True:

        kd = "Kids"

  else:
        kd = ''

  if isChecked_C5() == True:

        wm = "Women"
  else:
        wm = ''

  if isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == True:

   cats = bf + "|" + rd + "|" + jr + "|" + kd + "|" + wm
   success_inscrip()

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == False:

   cats = bf + "|" + rd + "|" + jr + "|" + kd

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == False:

   cats = bf + "|" + rd + "|" + jr

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == False and isChecked_C4() == False and isChecked_C5() == False:

   cats = bf + "|" + rd

  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == False and isChecked_C5() == False:

   cats = bf

  elif isChecked_C1() == False and isChecked_C2() == True and isChecked_C3() == False and isChecked_C4() == False and isChecked_C5() == False:

   cats = rd

  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == False:

   cats = jr

  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == False:

   cats = kd

  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == False:

   cats = wm

  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == False:

   cats = bf + "|" + jr

  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == False:

   cats = bf + "|" + kd

  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == False and isChecked_C5() == True:

   cats = bf + "|" + wm

  elif isChecked_C1() == False and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == False:

   cats = rd + "|" + jr

  elif isChecked_C1() == False and isChecked_C2() == True and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == False:

   cats = rd + "|"  + kd

  elif isChecked_C1() == False and isChecked_C2() == True and isChecked_C3() == False and isChecked_C4() == False and isChecked_C5() == True:

   cats = rd + "|"  + wm

  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == False:

   cats = jr + "|" + kd
  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == True:

   cats = jr + "|" + wm

  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == True:

   cats = kd + "|" + wm

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == False:

   cats = bf + "|" + rd + "|" + kd

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == False and isChecked_C4() == False and isChecked_C5() == True:

   cats = bf + "|" + rd + "|" + wm

  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == False:

   cats = bf + "|" + jr + "|" + kd
  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == True:

   cats = bf + "|" + jr + "|" + wm

  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == True:

   cats = bf + "|" + kd + "|" + wm

  elif isChecked_C1() == False and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == False:

   cats = rd + "|" + jr + "|" + kd

  elif isChecked_C1() == False and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == True:

   cats = rd + "|" + jr + "|" + wm

  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == True:

   cats =jr + "|" + kd + "|" + wm

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == False:

   cats = bf + "|" + rd + "|" + jr + "|" + kd

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == False and isChecked_C5() == True:

   cats = bf + "|" + rd + "|" + jr + "|" + wm

  elif isChecked_C1() == False and isChecked_C2() == True and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == True:

   cats = rd + "|" + jr + "|" + kd + "|" + wm

  elif isChecked_C1() == True and isChecked_C2() == False and isChecked_C3() == True and isChecked_C4() == True and isChecked_C5() == True:

   cats = bf + "|" + jr + "|" + kd + "|" + wm

  elif isChecked_C1() == True and isChecked_C2() == True and isChecked_C3() == False and isChecked_C4() == True and isChecked_C5() == True:

   cats = bf + "|" + rd + "|" + kd + "|" + wm

  elif isChecked_C1() == False and isChecked_C2() == False and isChecked_C3() == False and isChecked_C4() == False and isChecked_C5() == False:

   messagebox.showinfo("Warning", "Please select atlease one category for information to be registered")

  if isChecked_C1() == True or isChecked_C2() == True or isChecked_C3() == True or isChecked_C4() == True or isChecked_C5() == True:

        if embarcation.get() != "" and captain.get() != "" and number.get() != "":

            phone_num_l = len(number.get())

            if phone_num_l == 8:

              ws.append({'A': ID, 'B': embarcation.get(), 'C': captain.get(),'D': cats, 'E': number.get()})
              wb.save('XXII_RIFT.xlsx')

              success_inscrip()

              embarcation.delete(0, END)
              captain.delete(0, END)
              number.delete(0, END)

            else:

              wrong_num_length()

        else:

         failed_inscrip()

button_register = tk.Button(frame, text="Register", font="times 14", command=register_inscrip)
button_register.place(relx=0.45, rely=0.82)

#Design config for leaderboard
frame1 = tk.Frame(root, bg='white')
frame1.place(relx=0.03, rely=0.47, relwidth=0.45, relheight = 0.5)

bg_label = tk.Label(frame1, text='Leaderboard', font="Arialnova 25")
bg_label.place(relwidth=1, relheight=0.1)
#frame for leaderboard info
frame2 = tk.Frame(frame1)
frame2.place(relx=.08, rely=0.3, relwidth=.85, relheight = 0.6)

btn_color = "#ffebcc"
font_lead = "Centaur 18"

dt_label = tk.Label(frame1)
dt_label.place(relwidth=0.2, relheight=0.1)
clock()

date = dt.datetime.now()
format_date = f"{date:%a, %b %d %Y}"
w = Label(frame1, text=format_date, font=("times", 12), anchor='e')
w.place(relx = .7, relwidth=0.25, relheight=0.1)

#Design config for fish registry

register_font = "times 14"

frame3 = tk.Frame(root, bg='white')
frame3.place(relx=0.52, rely=0.05, relwidth=0.45, relheight = 0.92)

rg_label = tk.Label(frame3, text='Fish Registration', font="Arialnova 25")
rg_label.place(relwidth=1, relheight=0.065)

boat_label = tk.Label(frame3, text='Enter Boat ID: ', font="times 14")
boat_label.place(relx=0.32, rely=0.1)
boat= tk.Entry(frame3, cursor="arrow", font=register_font)
boat.place(relx=0.49, rely=0.10, relheight=0.045, relwidth=0.2)

frame4 = tk.Frame(frame3)
frame4.place(relx=0.25, rely=0.15, relwidth=0.5, relheight = 0.20)

def excel_view():
   os.chmod('XXII_RIFT.xlsx', S_IWRITE)
   os.system("start EXCEL.EXE XXII_RIFT.xlsx")

def find_boat(search_str):

 wb = load_workbook('XXII_RIFT.xlsx')
 ws = wb['Inscription']

 num_ent = search_str

 for cell in ws['A']:

    if cell.value == num_ent:

          cell_boat = cell
          return(cell.row)

def clear_info():

  for widgets in frame4.winfo_children():
      widgets.destroy()

def info_fishermen():

    hookup_time_1 = clock()
    dayy_1 = format_date
    print(hookup_time_1)
    print(dayy_1)

    wb = load_workbook('XXII_RIFT.xlsx')
    ws = wb['Inscription']

    clear_info()

    val_ent = boat.get()
    val_ent_int= int(val_ent)

    rows = find_boat(val_ent_int)

    row = ws[rows]

    for cell in row:

          if cell.column == 2:

           boat_pr = cell.value
           printable_boat = "Boat Name: " + boat_pr
           cat_d = ttk.Label(frame4, text = printable_boat,
              font = ("Times New Roman", 14))
           cat_d.place(relx=0.15, rely=0.1)

          elif cell.column == 3:

           cap_pr = cell.value
           printable_cap = "Captain Name: " + cap_pr
           cat_d1 = ttk.Label(frame4, text = printable_cap,
              font = ("Times New Roman", 14))
           cat_d1.place(relx=0.15, rely=0.28)

          elif cell.column == 4:

           cat = cell.value
           printable_cat = "Categories: " + cat
           cat_d2 = ttk.Label(frame4, text = printable_cat,
              font = ("Times New Roman", 14))
           cat_d2.place(relx=0.15, rely=0.46)

          elif cell.column == 5:

           num_pr = cell.value
           printable_num = "Phone Number: " + str(num_pr)
           cat_d3 = ttk.Label(frame4, text = printable_num,
              font = ("Times New Roman", 14))
           cat_d3.place(relx=0.15, rely=0.64)

          else:

            continue

    return(hookup_time_1)

button8 = tk.Button(frame3, text="Verify", bg=btn_color, font=register_font, command=info_fishermen)
button8.place(relx=0.80, rely=0.20)

hookup_label = tk.Label(frame3, text='Hook up: ', font="times 14")
hookup_label.place(relx=0.35, rely=0.38)
hookup = tk.Entry(frame3, cursor="arrow", font=register_font)
hookup.place(relx=0.47, rely=0.38, relheight=0.045, relwidth=0.2)

real_label = tk.Label(frame3, text='Release: ', font="times 14")
real_label.place(relx=0.20, rely=0.75)
real = tk.Entry(frame3, cursor="arrow", font=register_font)
real.place(relx=0.305, rely=0.75, relheight=0.045, relwidth=0.2)

button9 = tk.Button(frame3, text="View Sheet", bg=btn_color, font=register_font, command=excel_view)
button9.place(relx=0.415, rely=0.83)

cat_db = ttk.Label(frame3, text = "Select Category",
              font = ("Times New Roman", 15))
cat_db.place(relx=0.15, rely=0.5)

n = tk.StringVar()
categories = ttk.Combobox(frame3, width = 10, textvariable = n,
font = ("Times New Roman", 15))

# Adding combobox drop down list
categories['values'] = ('Billfish','Rodeo','Junior','Kids', 'Women')
categories.place(relx=0.33, rely=0.5)

fish_db_lbl =ttk.Label(frame3, text = "Fish",
          font = ("Times New Roman", 15))
fish_db_lbl.place(relx=0.60, rely=0.5)

x = tk.StringVar()
fishname= tk.Entry(frame3, cursor="arrow", font=register_font)
pounds= tk.Entry(frame3, cursor="arrow", font=register_font)

def fish_group(value):

 val = categories.get()

 fishname.place(relx=0.70, rely=0.5, relheight=0.045, relwidth=0.2)

 if val == "Rodeo" or val == "Junior" or val == "Kids" or val == "Women":
        pounds_db_lbl =ttk.Label(frame3, text = "Pounds",
          font = ("Times New Roman", 15))
        pounds_db_lbl.place(relx=0.60, rely=0.55)
        pounds.place(relx=0.70, rely=0.55, relheight=0.045, relwidth=0.2)

def save_fish():

  fish = fishname.get()

  return(fish)

button6 = tk.Button(frame3, text="Save", bg=btn_color, font=register_font, command=save_fish)
button6.place(relx=0.66, rely=0.62)

categories.bind("<<ComboboxSelected>>", fish_group)

def isChecked_C6():

   if CheckVar6.get() == 1:
        estado = TRUE
   else:
         estado = FALSE

   return(estado)

CheckVar6 = tk.IntVar()
C6 = tk.Checkbutton(frame3, text = "Clean Release", command=isChecked_C6, bg = "white", font=register_font, variable = CheckVar6,
onvalue = 1, offvalue = 0, height=2, width = 20, anchor='w')
C6.place(relx=0.6, rely=0.7325)

def info_forcat():

    wb = load_workbook('XXII_RIFT.xlsx')
    ws = wb['Inscription']

    num = boat.get()
    val_ent1 = num
    val_ent_int1= int(val_ent1)

    rows = find_boat(val_ent_int1)
    info = ()
    row = ws[rows]

    for cell in row:

          if cell.column == 2:

           boat_pr = cell.value
           printable_boat = "Boat Name: " + boat_pr
           cat_d = ttk.Label(frame4, text = printable_boat,
              font = ("Times New Roman", 14))
           cat_d.place(relx=0.15, rely=0.1)

          elif cell.column == 3:

           cap_pr = cell.value
           printable_cap = "Captain Name: " + cap_pr
           cat_d1 = ttk.Label(frame4, text = printable_cap,
              font = ("Times New Roman", 14))
           cat_d1.place(relx=0.15, rely=0.28)

          else:

            continue

    info = boat_pr, cap_pr

    return(info)

def fish_register():

 boat_info, cap_info = info_forcat()

 ID_1 = boat.get()

 print(ID_1)
 cat_store = categories.get()

 print(cat_store)

 fish_caught = save_fish()

 pounds_fish = pounds.get()

 if isChecked_C6() == True:
  clean_relpts = 50
  Clean_r = 'Yes'

 else:
  clean_relpts = 0
  Clean_r = 'No'

 if cat_store == "Billfish":

   if fish_caught == "Blue Marlin":
        fish_points = 500

   elif fish_caught == "White Marlin":

       fish_points = 300

   elif fish_caught == "Sailfish" or  fish_caught == "Spearfish":

       fish_points = 200

   else:
       fish_points = 0

 if cat_store == "Rodeo":

   wb = load_workbook('XXII_RIFT.xlsx')
   ws = wb['Rodeo']

   if fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 20:

         fish_points = pounds_fish * 1

   else:
         fish_points = 0

 if cat_store == "Junior":

   wb = load_workbook('XXII_RIFT.xlsx')
   ws = wb['Junior']

   if fish_caught == "Barracuda" or fish_caught == "Mackerel" or fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 15:

     fish_points = pounds_fish * 1

   else:
         fish_points = 0

 if cat_store == "Kids" :

   wb = load_workbook('XXII_RIFT.xlsx')
   ws = wb['Kids']

   if fish_caught == "Barracuda" or fish_caught == "Mackerel" or fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 15:

     fish_points = pounds_fish * 1
   else:
         fish_points = 0

 if cat_store == "Women":

   wb = load_workbook('XXII_RIFT.xlsx')
   ws = wb['Women']

   if fish_caught == "Barracuda" or fish_caught == "Mackerel" or fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 15:

     fish_points = pounds_fish * 1
   else:
         fish_points = 0

 total_points = clean_relpts + fish_points

 release_time = clock()
 day = format_date
 print(release_time)
 print(day)

 hookup_time = info_fishermen()

 fish_caught = save_fish()

 print(clean_relpts)

 print(fish_caught)

 print(total_points)

 print(ID_1)

 if cat_store == 'Billfish':


  wb = load_workbook('XXII_RIFT.xlsx')
  ws = wb["Billfish"]
  ws.append({'A': Iterador_billfish(),'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': total_points})
  wb.save('XXII_RIFT.xlsx')

 elif cat_store == 'Rodeo':

  wb = load_workbook('XXII_RIFT.xlsx')
  ws = wb["Rodeo"]
  ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
  wb.save('XXII_RIFT.xlsx')

 elif cat_store == 'Junior':

  wb = load_workbook('XXII_RIFT.xlsx')
  ws = wb["Rodeo"]
  ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
  wb.save('XXII_RIFT.xlsx')

 elif cat_store == 'Kids':

  wb = load_workbook('XXII_RIFT.xlsx')
  ws = wb["Rodeo"]
  ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
  wb.save('XXII_RIFT.xlsx')

 elif cat_store == 'Women':

  wb = load_workbook('XXII_RIFT.xlsx')
  ws = wb["Rodeo"]
  ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
  wb.save('XXII_RIFT.xlsx')

def reg_add_catch():

      screen = tk.Tk()
      screen.title('Multiple Catch Registry')

      HEIGHT = 644
      WIDTH = 720

      canvas = tk.Canvas(screen, height=HEIGHT, width=WIDTH)
      canvas.pack()

      register_font = "times 14"


      def clear_info_1():

       for widgets in frame6.winfo_children():
        widgets.destroy()

      frame5 = tk.Frame(screen, bg='white')
      frame5.place(relwidth=1, relheight = 1)

      rg_label = tk.Label(frame5, text='Fish Registration', font="Arialnova 25")
      rg_label.place(relwidth=1, relheight=0.065)

      boat_label = tk.Label(frame5, text='Enter Boat ID: ', font="times 14")
      boat_label.place(relx=0.32, rely=0.1)
      boat= tk.Entry(frame5, cursor="arrow", font=register_font)
      boat.place(relx=0.49, rely=0.10, relheight=0.045, relwidth=0.2)

      frame6 = tk.Frame(frame5)
      frame6.place(relx=0.25, rely=0.15, relwidth=0.5, relheight = 0.20)

      def info_fishermen_1():

            wb = load_workbook('XXII_RIFT.xlsx')
            ws = wb['Inscription']

            clear_info_1()

            val_ent_1 = boat.get()
            val_ent_int_1= int(val_ent_1)

            rows = find_boat(val_ent_int_1)

            row = ws[rows]

            for cell in row:

                  if cell.column == 2:

                        boat_pr = cell.value
                        printable_boat = "Boat Name: " + boat_pr
                        cat_d = ttk.Label(frame6, text = printable_boat,
                              font = ("Times New Roman", 14))
                        cat_d.place(relx=0.15, rely=0.1)

                  elif cell.column == 3:

                        cap_pr = cell.value
                        printable_cap = "Captain Name: " + cap_pr
                        cat_d1 = ttk.Label(frame6, text = printable_cap,
                              font = ("Times New Roman", 14))
                        cat_d1.place(relx=0.15, rely=0.28)

                  elif cell.column == 4:

                        cat = cell.value
                        printable_cat = "Categories: " + cat
                        cat_d2 = ttk.Label(frame6, text = printable_cat,
                              font = ("Times New Roman", 14))
                        cat_d2.place(relx=0.15, rely=0.46)

                  elif cell.column == 5:

                        num_pr = cell.value
                        printable_num = "Phone Number: " + str(num_pr)
                        cat_d3 = ttk.Label(frame6, text = printable_num,
                              font = ("Times New Roman", 14))
                        cat_d3.place(relx=0.15, rely=0.64)

                  else:

                        continue

      button9= tk.Button(frame5, text="Verify", bg=btn_color, font=register_font, command=info_fishermen_1)
      button9.place(relx=0.80, rely=0.20)

      hookup_label = tk.Label(frame5, text='Hook up: ', font="times 14")
      hookup_label.place(relx=0.35, rely=0.38)
      hookup = tk.Entry(frame5, cursor="arrow", font=register_font)
      hookup.place(relx=0.47, rely=0.38, relheight=0.045, relwidth=0.2)

      real_label = tk.Label(frame5, text='Release: ', font="times 14")
      real_label.place(relx=0.20, rely=0.75)
      real = tk.Entry(frame5, cursor="arrow", font=register_font)
      real.place(relx=0.305, rely=0.75, relheight=0.045, relwidth=0.2)

      button9 = tk.Button(frame5, text="View Sheet", bg=btn_color, font=register_font, command=excel_view)
      button9.place(relx=0.415, rely=0.83)

      cat_db = ttk.Label(frame5, text = "Select Category",
                  font = ("Times New Roman", 15))
      cat_db.place(relx=0.15, rely=0.5)

      n = tk.StringVar()
      categories = ttk.Combobox(frame5, width = 10, textvariable = n,
      font = ("Times New Roman", 15))

      # Adding combobox drop down list
      categories['values'] = ('Billfish','Rodeo','Junior','Kids', 'Women')
      categories.place(relx=0.33, rely=0.5)

      fish_db_lbl =ttk.Label(frame5, text = "Fish",
            font = ("Times New Roman", 15))
      fish_db_lbl.place(relx=0.60, rely=0.5)

      x = tk.StringVar()
      fishname= tk.Entry(frame5, cursor="arrow", font=register_font)
      pounds= tk.Entry(frame5, cursor="arrow", font=register_font)

      def fish_group(value):

            val = categories.get()

            fishname.place(relx=0.70, rely=0.5, relheight=0.045, relwidth=0.2)

            if val == "Rodeo" or val == "Junior" or val == "Kids" or val == "Women":
                  pounds_db_lbl =ttk.Label(frame5, text = "Pounds",
                  font = ("Times New Roman", 15))
                  pounds_db_lbl.place(relx=0.60, rely=0.55)
                  pounds.place(relx=0.70, rely=0.55, relheight=0.045, relwidth=0.2)

      def save_fish_1():

            fish = fishname.get()

            return(fish)

      categories.bind("<<ComboboxSelected>>", fish_group)

      button10 = tk.Button(frame5, text="Save", bg=btn_color, font=register_font, command=save_fish_1)
      button10.place(relx=0.66, rely=0.62)

      def isChecked_C7():

            if CheckVar7.get() == 1:
               estado = TRUE
            else:
               estado = FALSE

            return(estado)

      CheckVar7 = tk.IntVar()
      C7 = tk.Checkbutton(frame5, text = "Clean Release", command=isChecked_C7, bg = "white", font=register_font, variable = CheckVar6,
      onvalue = 1, offvalue = 0, height=2, width = 20, anchor='w')
      C7.place(relx=0.6, rely=0.7325)

      def info_forcat():

            wb = load_workbook('XXII_RIFT.xlsx')
            ws = wb['Inscription']

            num = boat.get()
            val_ent1 = num
            val_ent_int1= int(val_ent1)

            rows = find_boat(val_ent_int1)
            info = ()
            row = ws[rows]

            for cell in row:

                  if cell.column == 2:

                        boat_pr = cell.value
                        printable_boat = "Boat Name: " + boat_pr
                        cat_d = ttk.Label(frame4, text = printable_boat,
                              font = ("Times New Roman", 14))
                        cat_d.place(relx=0.15, rely=0.1)

                  elif cell.column == 3:

                        cap_pr = cell.value
                        printable_cap = "Captain Name: " + cap_pr
                        cat_d1 = ttk.Label(frame4, text = printable_cap,
                              font = ("Times New Roman", 14))
                        cat_d1.place(relx=0.15, rely=0.28)

                  else:

                        continue

            info = boat_pr, cap_pr

            return(info)

      def fish_register():

            boat_info, cap_info = info_forcat()

            ID_1 = boat.get()

            print(ID_1)
            cat_store = categories.get()

            print(cat_store)

            fish_caught = save_fish()

            pounds_fish = pounds.get()

            if isChecked_C7() == True:
                  clean_relpts = 50
                  Clean_r = 'Yes'

            else:
                  clean_relpts = 0
                  Clean_r = 'No'

            if cat_store == "Billfish":

             if fish_caught == "Blue Marlin":
                  fish_points = 500

             elif fish_caught == "White Marlin":

                  fish_points = 300

             elif fish_caught == "Sailfish" or  fish_caught == "Spearfish":

                  fish_points = 200

             else:
                  fish_points = 0

            if cat_store == "Rodeo":

             wb = load_workbook('XXII_RIFT.xlsx')
             ws = wb['Rodeo']

             if fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 20:

                        fish_points = pounds_fish * 1

             else:
                        fish_points = 0

            if cat_store == "Junior":

             wb = load_workbook('XXII_RIFT.xlsx')
             ws = wb['Junior']

             if fish_caught == "Barracuda" or fish_caught == "Mackerel" or fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 15:

              fish_points = pounds_fish * 1

             else:
                  fish_points = 0

            if cat_store == "Kids" :

             wb = load_workbook('XXII_RIFT.xlsx')
             ws = wb['Kids']

             if fish_caught == "Barracuda" or fish_caught == "Mackerel" or fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 15:

                      fish_points = pounds_fish * 1
             else:
                  fish_points = 0

            if cat_store == "Women":

             wb = load_workbook('XXII_RIFT.xlsx')
             ws = wb['Women']

             if fish_caught == "Barracuda" or fish_caught == "Mackerel" or fish_caught == "King Fish" or fish_caught == "Wahoo" or fish_caught == "MahiMahi" or fish_caught == "TUna" and pounds >= 15:

               fish_points = pounds_fish * 1
             else:
                  fish_points = 0

            total_points = clean_relpts + fish_points

            #hookup_time = hookup.get()
            #print(hookup_time)

            #release_time = real.get()
            #print(release_time)

            fish_caught = save_fish()

            print(clean_relpts)

            print(fish_caught)

            print(total_points)

            print(ID_1)

            if cat_store == 'Billfish':


                  wb = load_workbook('XXII_RIFT.xlsx')
                  ws = wb["Billfish"]
                  ws.append({'A': Iterador_billfish(),'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': total_points})
                  wb.save('XXII_RIFT.xlsx')

            elif cat_store == 'Rodeo':

                 wb = load_workbook('XXII_RIFT.xlsx')
                 ws = wb["Rodeo"]
                 ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
                 wb.save('XXII_RIFT.xlsx')

            elif cat_store == 'Junior':

                  wb = load_workbook('XXII_RIFT.xlsx')
                  ws = wb["Rodeo"]
                  ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
                  wb.save('XXII_RIFT.xlsx')

            elif cat_store == 'Kids':

                  wb = load_workbook('XXII_RIFT.xlsx')
                  ws = wb["Rodeo"]
                  ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
                  wb.save('XXII_RIFT.xlsx')

            elif cat_store == 'Women':

                  wb = load_workbook('XXII_RIFT.xlsx')
                  ws = wb["Rodeo"]
                  ws.append({'B': ID_1, 'C': boat_info, 'D': cap_info,'E': hookup_time, 'F': fish_caught, 'G': release_time, 'H': Clean_r, 'I': pounds_fish, 'J': total_points})
                  wb.save('XXII_RIFT.xlsx')

      button11 = tk.Button(frame5, text="Discard", bg=btn_color, font=register_font)
      button11.place(relx=0.145, rely=0.83)

      button12 = tk.Button(frame5, text="Register", bg=btn_color, font=register_font, command=fish_register)
      button12.place(relx=0.735, rely=0.83)

      screen.resizable(False, False)
      screen.mainloop()

button8 = tk.Button(frame3, text="Register Additional Catch", command=reg_add_catch, font=register_font, bg=btn_color)
button8.place(relx=0.35, rely=0.93)

###################################################################################################################

def info_for_LB(val):

    wb = load_workbook('XXII_RIFT.xlsx')
    ws = wb['Inscription']

    num = val
    val_ent1 = num
    val_ent_int1= int(val_ent1)

    rows = find_boat(val_ent_int1)
    info = ()
    row = ws[rows]

    for cell in row:

          if cell.column == 2:

           boat_pr = cell.value
           printable_boat = "Boat Name: " + boat_pr


          elif cell.column == 3:

           cap_pr = cell.value
           printable_cap = "Captain Name: " + cap_pr

          else:

            continue

    info = boat_pr, cap_pr

    return(info)

def Billfish_LB_info():

      req_col = [0, 1]

      billfish = pd.read_excel('XXII_RIFT.xlsx',sheet_name='Billfish', index_col=0, usecols=req_col)
      ins_info = pd.read_excel('XXII_RIFT.xlsx',sheet_name='Inscription', index_col=0)
      billfish_all = pd.read_excel('XXII_RIFT.xlsx',sheet_name='Billfish', index_col=0)
      Id_col = billfish.filter(items=['ID'])

      i = Id_col.max()

      i_1 = int(i)

      print(i_1)

      itera = 1
      row_col_num = 2
      x = 1

      wb = load_workbook('XXII_RIFT.xlsx')
      ws = wb["Billfish_LB"]

      for row in ws['A2:E5']:
       for cell in row:
        cell.value = None

      while (x <= i_1):

       boat_1, cap_1 = info_for_LB(x)

       print(x)

       print(billfish_all[billfish_all['ID'] == x])
       data = billfish_all[billfish_all['ID'] == x]

       data_points = data['Points']
       #values_points = data_points.filter(items='Points')

       sum = str(data['Points'].sum())

       print(data_points)
       #print(values_points)
       print(sum)

       wb = load_workbook('XXII_RIFT.xlsx')
       ws = wb["Billfish_LB"]

       #sum_val = data.sum(axis=7)
       #print("Total Points" + str(sum_val))
       if sum == "0":

            x = x + 1
            continue
       else:
       #ws.append({'A': str(itera),'B': str(x), 'C': boat_1, 'D': cap_1,'E': sum})
       #wb.save('XXII_RIFT.xlsx')

        wb = load_workbook('XXII_RIFT.xlsx')
        ws = wb["Billfish_LB_sorted"]

        ws["A"+str(row_col_num)]= itera
        ws["B"+str(row_col_num)]= str(x)
        ws["C"+str(row_col_num)]= boat_1
        ws["D"+str(row_col_num)]= cap_1
        ws["E"+str(row_col_num)]= sum

        wb.save('XXII_RIFT.xlsx')

        row_col_num = row_col_num + 1
        itera = itera + 1
        x = x + 1


      wb = load_workbook('XXII_RIFT.xlsx')
      ws = wb["Billfish_LB_sorted"]

      x_len = str(len(ws['A']))
      start = "A2"
      end = "E"+ x_len

      print(start)
      print(end)

      print(x_len)

      billfish_LB_info = pd.read_excel('XXII_RIFT.xlsx',sheet_name='Billfish_LB', index_col=0)
      print(billfish_LB_info)
      sorted_BLB = billfish_LB_info.sort_values(by="Total Points", ascending=False)

      group = start + ":" + end

      print(group)

      s_BLB = pd.DataFrame(sorted_BLB)

      print(s_BLB)

      return(s_BLB)

def leaderboard_billfish():

      Data = Billfish_LB_info()
      dt_label_1 = tk.Label(frame2)
      dt_label_1.place(relwidth=0.65, relheight=0.65)

      dt_label_1.config(text=Data,font='times 16')

button = tk.Button(frame1, text="Billfish",command=leaderboard_billfish, font=font_lead, bg=btn_color)
button.place(relx=0.15, rely=0.15)

button1 = tk.Button(frame1, text="Rodeo", bg=btn_color, font=font_lead)
button1.place(relx=0.30, rely=0.15)

button2 = tk.Button(frame1, text="Junior", bg=btn_color, font=font_lead)
button2.place(relx=0.45, rely=0.15)

button3 = tk.Button(frame1, text="Kids", bg=btn_color, font=font_lead)
button3.place(relx=0.60, rely=0.15)

button4 = tk.Button(frame1, text="Women", bg=btn_color, font=font_lead)
button4.place(relx=0.75, rely=0.15)

button5 = tk.Button(frame3, text="Discard", bg=btn_color, font=register_font)
button5.place(relx=0.145, rely=0.83)

button7 = tk.Button(frame3, text="Register", bg=btn_color, font=register_font, command=fish_register)
button7.place(relx=0.735, rely=0.83)

root.resizable(False, False)
root.mainloop()


