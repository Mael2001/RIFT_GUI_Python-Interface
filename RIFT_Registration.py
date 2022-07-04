#Base Libraries
from cgitb import text
from logging import PlaceHolder
from wsgiref import validate
import pandas as pd
import os
from openpyxl import *
from datetime import datetime
import time
import webbrowser

#Graphic Libraries
import tkinter.messagebox
import tkinter
import customtkinter
import threading

#Image Libraries

#SIZE
COMPLETE_WIDTH = 700
COMPLETE_HEIGHT = 800

#Working Directories
CURRENT_DIRECTORY = os.path.dirname(os.path.realpath(__file__))
FILE_NAME=".\dump\\test.xlsx"
FILE_PATH = os.path.join(CURRENT_DIRECTORY,FILE_NAME)

#Base Settings
customtkinter.set_appearance_mode("system")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green

class Registration(customtkinter.CTk):

    APP_WIDTH = COMPLETE_WIDTH
    APP_HEIGHT = COMPLETE_HEIGHT
    current_team=[]

    def __init__(self):
        super().__init__()
        WIDTH = self.winfo_screenwidth()
        HEIGHT = self.winfo_screenwidth()

        app_center_coordinate_x = (WIDTH/2) - (Registration.APP_WIDTH )
        app_center_coordinate_y = (HEIGHT/2) - (Registration.APP_HEIGHT)

        self.title("Plataforma RIFT")
        self.geometry(f"{Registration.APP_WIDTH}x{Registration.APP_HEIGHT}+{int(app_center_coordinate_x)}+{int(app_center_coordinate_y)}")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)  # call .on_closing() when app gets closed
        self.configure(fg_color=("#189FE7"))

        #GUI TITLE
        self.homeLabel = customtkinter.CTkLabel(master=self,
                                                    corner_radius=7,
                                                    height=40,
                                                    fg_color=("white", "grey38"),  # <- custom tuple-color
                                                    text_font=("Arial",14),
                                                    text="Hookup Registration")
        self.homeLabel.grid(row=0,column=0,columnspan=2,padx=15,pady=10)

        #BASE FRAME
        self.frame_home = customtkinter.CTkFrame(master=self,
                                                    width= Registration.APP_WIDTH,
                                                    height= Registration.APP_HEIGHT,
                                                    fg_color="black",
                                                    corner_radius=10)
        self.frame_home.grid(row=1,column=0,padx=15, pady=20,sticky="nsew")


        #BOAT NAME
        self.boat_name = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Enter Boat ID: ")
        self.boat_name.grid(row=0,column=0,padx=10,pady=20)

        #BOAT INPUT
        self.boat_input= customtkinter.CTkEntry(master=self.frame_home,
                                            width=Registration.APP_WIDTH/2+50,
                                            placeholder_text="Boat ID")
        self.boat_input.grid(row=0,column=1, columnspan=2, pady=20, padx=20, sticky="we")

        #SEARCH BUTTON
        self.search_button= customtkinter.CTkButton(master=self.frame_home,
                                            width=Registration.APP_WIDTH/2-50,
                                            text="Search",
                                            command=self.search)
        self.search_button.grid(row=1,column=0,columnspan=3,pady=20, padx=20, sticky="we")

        #SEARCH FRAME
        self.search_frame = customtkinter.CTkFrame(master=self.frame_home,
                                                    width= Registration.APP_WIDTH,
                                                    height= Registration.APP_HEIGHT,
                                                    fg_color="white",
                                                    corner_radius=10)
        self.search_frame.grid(row=2,column=0,columnspan=4,padx=15, pady=20,sticky="nsew")

        #BOAT NAME
        self.search_boat_name = customtkinter.CTkEntry(master=self.search_frame,
                                                    corner_radius=7,
                                                    width= Registration.APP_WIDTH/2-50,
                                                    placeholder_text="Boat Name",
                                                    text="Boat")
        self.search_boat_name.grid(row=0,column=0,columnspan=2,padx=10,pady=20)
        self.search_boat_name.configure(state=tkinter.DISABLED)

        #CAPTAIN NAME
        self.search_captain_name = customtkinter.CTkEntry(master=self.search_frame,
                                                    corner_radius=7,
                                                    width= Registration.APP_WIDTH/2-50,
                                                    placeholder_text="Captain Name",
                                                    text="Captain")
        self.search_captain_name.grid(row=0,column=3,columnspan=2,padx=10,pady=20)
        self.search_captain_name.configure(state=tkinter.DISABLED)

        #CATEGORY SELECT LABEL
        self.category_combobox_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Select Category: ")
        self.category_combobox_label.grid(row=3,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #CATEGORY SELECT
        self.category_combobox = customtkinter.CTkOptionMenu(master=self.frame_home,
                                                    values=["Billfish", "Rodeo", "Junior","Kids", "Women", "None"])
        self.category_combobox.grid(row=3,column=1,columnspan=2,padx=15, pady=20,sticky="nsew")
        self.category_combobox.set("None")
        self.category_combobox.configure(state=tkinter.DISABLED)


        #FISH SELECT LABEL
        self.fish_combobox_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Select Fish: ")
        self.fish_combobox_label.grid(row=4,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #FISH SELECT
        self.fish_combobox = customtkinter.CTkOptionMenu(master=self.frame_home,
                                                    values=["Fish", "None"])
        self.fish_combobox.grid(row=4,column=1,columnspan=2,padx=15, pady=20,sticky="nsew")
        self.fish_combobox.set("None")
        self.fish_combobox.configure(state=tkinter.DISABLED)

        #HOOKUP TIME LABEL
        self.hookup_time_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Hookup Time: ")
        self.hookup_time_label.grid(row=5,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #HOOKUP TIME
        self.hookup_time = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text=datetime.now().strftime("%H:%M:%S"),
                                                    width= Registration.APP_WIDTH/2-50)
        self.hookup_time.grid(row=5,column=1,columnspan=1,padx=15, pady=20,sticky="nsew")

        #REGISTER HOOKUP
        self.hookup_button = customtkinter.CTkButton(master=self.frame_home,
                                            width=Registration.APP_WIDTH/2-50,
                                            text="Register Hookup",
                                            command=self.register_hookup)
        self.hookup_button.grid(row=6,column=0,columnspan=3,padx=15, pady=20,sticky="nsew")
        self.hookup_button.configure(state=tkinter.DISABLED)

        #RELEASE TIME LABEL
        self.release_time_label = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text="Release Time: ")
        self.release_time_label.grid(row=7,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #RELEASE TIME
        self.release_time = customtkinter.CTkLabel(master=self.frame_home,
                                                    corner_radius=7,
                                                    text=datetime.now().strftime("%H:%M:%S"),
                                                    width= Registration.APP_WIDTH/2-50)
        self.release_time.grid(row=7,column=1,columnspan=1,padx=15, pady=20,sticky="nsew")

        #CLEAN RELEASE CHECK
        self.clean_release=customtkinter.CTkCheckBox(master=self.frame_home,
                                                    text="Clean Release")
        self.clean_release.grid(row=7,column=2,columnspan=1,padx=15, pady=20,sticky="nsew")
        self.clean_release.configure(state=tkinter.DISABLED)

        #DISCARD BUTTON
        self.discard_button = customtkinter.CTkButton(master=self.frame_home,
                                            text="Discard",
                                            command=self.discard)
        self.discard_button.grid(row=8,column=0,columnspan=1,padx=15, pady=20,sticky="nsew")

        #CHECK XLSX BUTTON
        self.discard_button = customtkinter.CTkButton(master=self.frame_home,
                                            text="View Worksheet",
                                            command=self.view)
        self.discard_button.grid(row=8,column=1,columnspan=1,padx=15, pady=20,sticky="nsew")

        #REGISTER RELEASE
        self.release_button = customtkinter.CTkButton(master=self.frame_home,
                                            text="Register Release",
                                            command=self.register_release)
        self.release_button.grid(row=8,column=2,columnspan=1,padx=15, pady=20,sticky="nsew")
        self.release_button.configure(state=tkinter.DISABLED)

    #Register hookup
    def register_hookup(self):
        if(self.validate_inputs()):
            sheet_name=self.category_combobox.get()
            wb = load_workbook(filename=FILE_NAME)
            ws = wb[sheet_name]
            if(ws.max_row < 1):
                headers = [
                    "ID",
                    "BOAT_NAME",
                    "CAPTAIN_NAME",
                    "HOOKUP_TIME",
                    "FISH",
                    "RELEASE_TIME",
                    "CLEAN_RELEASE",
                    "POINTS"
                    ]
            self.write_to_sheet(FILE_NAME,sheet_name,headers)
        else:
            self.trigger_error("Please Fix Errors in the form","Unable to Continue")
            return
        sheet_name=self.category_combobox.get()
        data = [
            self.current_team['ID'],
            self.current_team['BOAT_NAME'],
            self.current_team['CAPTAIN_NAME'],
            self.hookup_time.get(),
            self.fish_combobox.get(),
            "None",
            "No",
            50
        ]
        print(f"Wrote to {FILE_NAME}, sheet {sheet_name}")
        self.write_to_sheet(FILE_NAME,sheet_name,data)

    #Validation for all inputs
    def validate_inputs(self):
        if(self.category_combobox.get() == "None" or self.current_team[self.category_combobox.get()] == "Not Participating"):
            self.trigger_error("Team is not participating in this Category", "Input error")
            return False
        if(self.fish_combobox.get() == "None"):
            self.trigger_error("Invalid Fish Selected", "Input error")
            return False
        return True
    #Register release
    def register_release(self):
        if(self.validate_inputs()):
            sheetName=self.category_combobox.get()
            hookup_result = self.search_hookup(sheetName)
            self.rewrite_row(FILE_NAME,sheetName,hookup_result)
            print(f"Wrote to {FILE_NAME}, sheet {sheetName}")
        else:
            self.trigger_error("Please Fix Errors in the form","Unable to Continue")
            return

    #Search Hookup Information
    def search_hookup(self,sheetName):
        excel_data = pd.read_excel(FILE_NAME,sheet_name=sheetName)
        boat_id = self.boat_input.get()
        hookup = excel_data.iloc[excel_data['ID'] == int(boat_id)]
        return hookup.iloc[-1]

    def rewrite_row(self,path,sheetName,hookup_result):
        wb = load_workbook(path)
        ws = wb[sheetName]
        row = hookup_result['ID']+1
        RELEASE_TIME = ws.cell(row=row,column=6)
        RELEASE_TIME.value = self.release_time.get()
        CLEAN_RELEASE = ws.cell(row=row,column=7)
        CLEAN_RELEASE.value = "Yes" if self.clean_release.get() else "No"
        POINTS = ws.cell(row=row,column=8)
        POINTS.value = 550
        wb.save(path)

    #View XLSX file
    def view(self):
        webbrowser.open(FILE_NAME)
    #Discard all changes
    def discard(self):
        #Resetting Search Results
        self.search_boat_name.configure(state=tkinter.NORMAL)
        self.search_captain_name.configure(state=tkinter.NORMAL)
        self.search_boat_name.configure(placeholder_text="Boat Name")
        self.search_captain_name.configure(placeholder_text="Captain Name")
        self.search_boat_name.configure(state=tkinter.DISABLED)
        self.search_captain_name.configure(state=tkinter.DISABLED)
        #Refreshing Clocks
        self.hookup_time.configure(text=datetime.now().strftime("%H:%M:%S"))
        self.release_time.configure(text=datetime.now().strftime("%H:%M:%S"))
        #Resetting default values
        self.boat_input.configure(text="")
        self.category_combobox.set("None")
        self.fish_combobox.set("None")
        self.clean_release.deselect()
        #Disabling all controls when refreshing form
        self.category_combobox.configure(state=tkinter.DISABLED)
        self.fish_combobox.configure(state=tkinter.DISABLED)
        self.hookup_button.configure(state=tkinter.DISABLED)
        self.clean_release.configure(state=tkinter.DISABLED)
        self.release_button.configure(state=tkinter.DISABLED)
        return True
    #Search ID
    def search(self):
        excel_data = pd.read_excel(FILE_NAME,sheet_name="Inscription")
        boat_id = self.boat_input.get()
        for result in (excel_data.isin([int(boat_id)]).any()):
            if result:
                self.current_team = excel_data.iloc[int(boat_id)-1]
                #Writing Search Results
                self.search_boat_name.configure(state=tkinter.NORMAL)
                self.search_captain_name.configure(state=tkinter.NORMAL)
                self.search_boat_name.configure(placeholder_text=self.current_team['BOAT_NAME'])
                self.search_captain_name.configure(placeholder_text=self.current_team['CAPTAIN_NAME'])
                self.search_boat_name.configure(state=tkinter.DISABLED)
                self.search_captain_name.configure(state=tkinter.DISABLED)
                #Enabling all controls when Successfully found ID
                self.category_combobox.configure(state=tkinter.NORMAL)
                self.fish_combobox.configure(state=tkinter.NORMAL)
                self.hookup_button.configure(state=tkinter.NORMAL)
                self.clean_release.configure(state=tkinter.NORMAL)
                self.release_button.configure(state=tkinter.NORMAL)
        self.trigger_error(f"No Inscription was found matching {boat_id}","No Register Found")

    #Writing to File
    def write_to_sheet(self,path,sheetName,data):
        wb = load_workbook(path)
        ws = wb[sheetName]
        ws.append(data)
        wb.save(path)
    #Update HookupTime
    def updateClock(self):
        time.sleep(1)
        while True:
            self.hookup_time.configure(text=datetime.now().strftime("%H:%M:%S"))
            self.release_time.configure(text=datetime.now().strftime("%H:%M:%S"))
            time.sleep(1)
    #Error Box Trigger
    def trigger_error(self,ErrorMsg,Error):
        tkinter.messagebox.showerror(title=Error, message=ErrorMsg)
    #Close Program
    def on_closing(self, event=0):
        self.destroy()